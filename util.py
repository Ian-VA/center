import json
import math
from functools import lru_cache
from pathlib import Path
import openpyxl
import pandas as pd
from data import CobraData
from pollution import datacenter_pollution, DEFAULT_UTILIZATION, DEFAULT_PUE

ROOT = Path(__file__).parent

@lru_cache(maxsize=1)
def _load_counties() -> list[dict]:
    """Load counties data from JSON file."""
    counties_path = ROOT / "data" / "counties.json"
    with open(counties_path, "r") as f:
        return json.load(f)

def _haversine_miles(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Calculate the great-circle distance between two points on Earth in miles."""
    R = 3959.0  # Earth's radius in miles
    lat1r, lat2r = math.radians(lat1), math.radians(lat2)
    dlat, dlon = math.radians(lat2 - lat1), math.radians(lon2 - lon1)
    a = math.sin(dlat / 2) ** 2 + math.cos(lat1r) * math.cos(lat2r) * math.sin(dlon / 2) ** 2
    return 2 * R * math.asin(math.sqrt(a))

def estimate_county(lat: float, lon: float) -> dict:
    """Estimate the county for given lat/lon by finding the closest county centroid."""
    counties = _load_counties()
    best_distance = float("inf")
    best_county = None
    for county in counties:
        clat = county.get("LAT")
        clon = county.get("LON")
        if clat is None or clon is None:
            continue
        distance = _haversine_miles(lat, lon, clat, clon)
        if distance < best_distance:
            best_distance = distance
            best_county = county
    if best_county is None:
        raise ValueError(f"No county found for lat={lat}, lon={lon}")
    return {
        "state": best_county["STNAME"],
        "county": best_county["CYNAME"],
        "fips": best_county["FIPS"],
        "distance_miles": round(best_distance, 2)
    }


EGRID_PATH = ROOT / "center" / "echo_bulk" / "egrid2023.xlsx"
EGRID_URL = "https://www.epa.gov/system/files/documents/2025-01/egrid2023_data_rev1.xlsx"

_RESOURCE_MIX_CODES = {
    "coal": "SRCLPR", "oil": "SROLPR", "gas": "SRGSPR", "nuclear": "SRNCPR",
    "hydro": "SRHYPR", "biomass": "SRBMPR", "wind": "SRWIPR", "solar": "SRSOPR",
    "geothermal": "SRGTPR", "other_fossil": "SROFPR",
    "total_nonrenewable": "SRTNPR", "total_renewable": "SRTRPR",
}

@lru_cache(maxsize=1)
def _open_egrid() -> openpyxl.Workbook:
    if not EGRID_PATH.exists():
        import sys
        import urllib.request
        EGRID_PATH.parent.mkdir(parents=True, exist_ok=True)
        print(f"[eGRID] file missing — downloading {EGRID_URL} → {EGRID_PATH} (~21MB)",
              file=sys.stderr)
        urllib.request.urlretrieve(EGRID_URL, EGRID_PATH)
        print(f"[eGRID] saved {EGRID_PATH.stat().st_size/1e6:.0f} MB", file=sys.stderr)
    return openpyxl.load_workbook(str(EGRID_PATH), read_only=True, data_only=True)

@lru_cache(maxsize=1)
def _subregion_table() -> dict:
    """{SUBRGN: {col_code: value, ...}} from SRL23 sheet."""
    wb = _open_egrid()
    ws = wb["SRL23"]
    rows = list(ws.iter_rows(values_only=True))
    codes = rows[1]
    out = {}
    for r in rows[2:]:
        if not r or not r[1]:
            continue
        d = {c: v for c, v in zip(codes, r) if c}
        sub = d.get("SUBRGN")
        if isinstance(sub, str):
            out[sub] = d
    return out

@lru_cache(maxsize=1)
def _plant_locations() -> list[tuple]:
    """List of (lat, lon, subrgn) for every eGRID plant — used for nearest-plant lookup."""
    wb = _open_egrid()
    ws = wb["PLNT23"]
    rows = list(ws.iter_rows(values_only=True))
    codes = rows[1]
    idx_lat = codes.index("LAT")
    idx_lon = codes.index("LON")
    idx_sub = codes.index("SUBRGN")
    out = []
    for r in rows[2:]:
        if not r:
            continue
        lat = r[idx_lat]; lon = r[idx_lon]; sub = r[idx_sub]
        if isinstance(lat, (int, float)) and isinstance(lon, (int, float)) and isinstance(sub, str):
            if math.isfinite(lat) and math.isfinite(lon):
                out.append((lat, lon, sub))
    return out

def lookup_subregion(lat: float, lon: float) -> str:
    """Find eGRID subregion code via the nearest power plant in the eGRID plant list."""
    plants = _plant_locations()
    best_d = float("inf"); best = None
    for plat, plon, sub in plants:
        d = _haversine_miles(lat, lon, plat, plon)
        if d < best_d:
            best_d = d; best = sub
    if best is None:
        raise ValueError(f"No eGRID plant found near lat={lat}, lon={lon}")
    return best

def most_common_fuel(lat: float, lon: float) -> str:
    """Compute the most common fuel used for electricity generation in the EPA subregion for given coordinates."""
    sub_code = lookup_subregion(lat, lon)
    sub = _subregion_table().get(sub_code)
    if not sub:
        raise ValueError(f"No data for subregion {sub_code}")
    
    mix = {}
    for fuel, code in _RESOURCE_MIX_CODES.items():
        if "total_" not in fuel:  # Exclude totals
            v = sub.get(code)
            mix[fuel] = v if isinstance(v, (int, float)) else 0.0
    
    if not mix:
        raise ValueError(f"No resource mix data for subregion {sub_code}")
    
    most_common = max(mix, key=mix.get)
    return most_common

# EPA Tier 4 final / NSPS-compliant stationary diesel (>560 kW), lb/MMBtu.
# NOx ~0.50 (≈3.5 g/bhp-hr after SCR), ULSD-grade SO2,
# PM2.5 ~0.005 (Tier 4 PM standard ~0.04 g/bhp-hr after DPF), VOC ~0.014.
DIESEL_EMISSION_FACTOR = {
    "NOx": 0.50,
    "SO2": 0.029,
    "PM25": 0.005,
    "VOC": 0.014,
}  # lb/MMBtu
GAS_EMISSION_FACTOR = {
    "NOx": 1.94,
    "SO2": 5.88e-4,
    "VOC": 1.2e-1,
    "PM25": 3.84e-2,
}  # lb/MMBtu
MWH_TO_MMBTU = 3.41
LB_TO_KG = 0.453592
LB_PER_TON = 2000.0  # US short ton — matches grid emissions and the COBRA emissions table


def _annual_mwh_for_generator(
    power_mw: float,
    mode: str,
    run_hours: float,
    utilization: float = DEFAULT_UTILIZATION,
    pue: float = DEFAULT_PUE,
) -> float:
    """Annual MWh delivered by a single generator.

    - Prime: continuous, mw × 8760 × util × pue.
    - Backup: only the hours it actually runs, mw × run_hours.
    """
    if mode == "prime":
        return power_mw * 8760.0 * utilization * pue
    return max(0.0, power_mw) * max(0.0, run_hours)


def generator_pollution(
    power_mw: float,
    is_diesel: bool,
    mode: str = "prime",
    run_hours: float = 0.0,
    utilization: float = DEFAULT_UTILIZATION,
    pue: float = DEFAULT_PUE,
):
    annual_MWh = _annual_mwh_for_generator(power_mw, mode, run_hours, utilization, pue)
    annual_MMBtu = annual_MWh * MWH_TO_MMBTU
    factor = DIESEL_EMISSION_FACTOR if is_diesel else GAS_EMISSION_FACTOR
    # Short tons/year — matches grid_emissions and the COBRA table.
    return {k: v * annual_MMBtu / LB_PER_TON for k, v in factor.items()}


def load_sectors() -> pd.DataFrame:
    return pd.read_json(ROOT / "data" / "sectors.json")


def get_grid_fuel_id(common_fuel: str) -> int:
    if common_fuel == "coal":
        return 544  # Bit coal
    elif common_fuel == "gas":
        return 545  # Natural gas
    else:
        return 548  # Distillate oil


def get_generator_fuel_id(is_diesel: bool) -> int:
    if is_diesel:
        return 548
    else:
        return 545  # Natural Gas


def get_tier_ids(sectors: pd.DataFrame, fuel_id: int) -> list:
    return list(
        sectors[sectors["ID"] == fuel_id][
            ["TIER1", "TIER2", "TIER3"]
        ].iloc[0]
    )


def load_counties() -> pd.DataFrame:
    return pd.read_json(ROOT / "data" / "counties.json")


def get_source_index(counties: pd.DataFrame, fips: str) -> int:
    return counties[counties["FIPS"] == int(fips)]["SOURCEINDX"].iloc[0]


_COBRA_POLLUTANTS = ['NOx', 'SO2', 'NH3', 'SOA', 'PM25', 'VOC']


@lru_cache(maxsize=None)
def _grid_sourceindxs_for_subregion(subregion_code: str) -> tuple:
    """Counties (by SOURCEINDX) that contain power plants in the given eGRID subregion.

    Grid emissions get added at the *generating sources*, not at the user's data-center
    county, because the data center pulls electrons from these plants — not burns fuel
    onsite. Cached so the eGRID plant scan runs once per subregion.
    """
    counties = load_counties()
    indices = set()
    for plat, plon, sub in _plant_locations():
        if sub != subregion_code:
            continue
        diffs = (counties['LAT'] - plat) ** 2 + (counties['LON'] - plon) ** 2
        idx = int(diffs.idxmin())
        indices.add(int(counties.loc[idx, 'SOURCEINDX']))
    return tuple(sorted(indices))


def _add_to_baseline(
    data: CobraData,
    raw_base,
    raw_modified,
    addition: dict,
    tier_ids: list,
    sourceindxs: list,
):
    """Add `addition` (tons/yr per pollutant) on top of baseline for matching sector/counties.

    CobraData.modify_emissions uses REPLACE semantics (sets new total = payload). For a
    new-facility scenario we want ADD: pass `existing_baseline + addition` as the new
    total. This also preserves baseline for pollutants the addition doesn't mention,
    instead of accidentally zeroing PM25/VOC/NH3 because the payload happened to omit them.
    """
    if not any(addition.get(p, 0) for p in _COBRA_POLLUTANTS):
        return raw_modified
    mask = (
        (raw_base['TIER1'] == int(tier_ids[0])) &
        (raw_base['TIER2'] == int(tier_ids[1])) &
        (raw_base['TIER3'] == int(tier_ids[2])) &
        (raw_base['sourceindx'].isin(sourceindxs))
    )
    existing = raw_base.loc[mask, _COBRA_POLLUTANTS].sum()
    new_total = {p: float(existing.get(p, 0)) + float(addition.get(p, 0)) for p in _COBRA_POLLUTANTS}
    return data.modify_emissions(raw_modified, new_total, tier_ids, sourceindxs)


def prepare_emissions(
    lat: float,
    lon: float,
    total_power: float,
    generators: list | None = None,
    *,
    generator_power: float = 0.0,
    generator_is_diesel: bool = True,
    utilization: float = DEFAULT_UTILIZATION,
    pue: float = DEFAULT_PUE,
) -> tuple:
    """Build base + modified emission tables for COBRA.

    `generators` is a list of dicts {fuel, powerMW, mode, runHours}. When omitted the
    legacy single-generator params are used (assumed prime power) for backwards compat.
    """
    sectors = load_sectors()
    county = estimate_county(lat, lon)
    counties = load_counties()
    site_sourceindx = get_source_index(counties, county["fips"])
    data = CobraData()
    raw_base = data.load_emissions_base()
    base_emissions = data.summarize_emissions(raw_base)

    diesel_total: dict = {}
    gas_total: dict = {}
    gen_annual_MWh_total = 0.0

    if generators is None:
        generators = [{
            "fuel": "Diesel" if generator_is_diesel else "Natural Gas",
            "powerMW": generator_power,
            "mode": "prime",
            "runHours": 0.0,
        }]

    for gen in generators:
        power_mw = float(gen.get("powerMW", 0) or 0)
        if power_mw <= 0:
            continue
        is_diesel = gen.get("fuel") == "Diesel"
        mode = gen.get("mode", "prime")
        run_hours = float(gen.get("runHours", 0) or 0)
        gen_annual_MWh_total += _annual_mwh_for_generator(
            power_mw, mode, run_hours, utilization, pue
        )
        emissions = generator_pollution(
            power_mw, is_diesel, mode=mode, run_hours=run_hours,
            utilization=utilization, pue=pue,
        )
        bucket = diesel_total if is_diesel else gas_total
        for k, v in emissions.items():
            bucket[k] = bucket.get(k, 0.0) + v

    # Grid energy = facility load minus what generators actually deliver. For backup
    # gens that almost never run, gen energy ≈ 0 → grid_mw ≈ total_power, which is
    # what you'd expect (the data center is grid-fed almost all the time).
    facility_annual_MWh = total_power * 8760.0 * utilization * pue
    grid_annual_MWh = max(0.0, facility_annual_MWh - gen_annual_MWh_total)
    grid_equivalent_mw = (
        grid_annual_MWh / (8760.0 * utilization * pue) if utilization * pue > 0 else 0.0
    )

    grid_meta = datacenter_pollution(
        lat, lon, grid_equivalent_mw, utilization=utilization, pue=pue
    )
    common_fuel = most_common_fuel(lat, lon)
    grid_fuel_id = get_grid_fuel_id(common_fuel)
    grid_tier_ids = get_tier_ids(sectors, grid_fuel_id)
    grid_emissions = {
        k: (v["tons_per_year"] if isinstance(v, dict) else 0) or 0
        for k, v in grid_meta["emissions"].items()
    }

    # Distribute grid emissions across the eGRID subregion's *generating plants*.
    subregion_code = grid_meta["geo"]["subregion_code"]
    plant_sourceindxs = list(_grid_sourceindxs_for_subregion(subregion_code))
    if not plant_sourceindxs:
        plant_sourceindxs = [site_sourceindx]

    modified = _add_to_baseline(
        data, raw_base, raw_base, grid_emissions, grid_tier_ids, plant_sourceindxs
    )

    # Onsite generators are physically at the user's site.
    if any(diesel_total.values()):
        diesel_tier = get_tier_ids(sectors, get_generator_fuel_id(True))
        modified = _add_to_baseline(
            data, raw_base, modified, diesel_total, diesel_tier, [site_sourceindx]
        )
    if any(gas_total.values()):
        gas_tier = get_tier_ids(sectors, get_generator_fuel_id(False))
        modified = _add_to_baseline(
            data, raw_base, modified, gas_total, gas_tier, [site_sourceindx]
        )

    modified_emissions = data.summarize_emissions(modified)
    return data, base_emissions, modified_emissions
