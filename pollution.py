"""eGRID-based grid pollution conversion for US data centers.

    datacenter_pollution(lat, lon, mw_capacity) -> dict

Translates a data center's MW capacity into annual grid-import pollution using EPA eGRID 2023
subregion emission factors. Returns pounds and short-tons for every pollutant eGRID tracks
(CO2, NOx, SO2, CH4, N2O, Hg, CO2-equivalent), plus an EPA-social-cost USD/year monetization.

Pipeline:
    1. annual_MWh = mw_capacity × 8760 × utilization × PUE
    2. lat/lon -> nearest eGRID power plant -> its subregion (RFCE, ERCT, etc.)
    3. emissions = annual_MWh × subregion_factor (lb/MWh) / 2000 -> short tons/year
    4. social_cost = sum(tons_pollutant × EPA $/ton)

eGRID column codes used (from the SRL23 sheet):
    Average rates ("total output"):  SRNOXRTA, SRSO2RTA, SRCO2RTA, SRCH4RTA, SRN2ORTA, SRC2ERTA, SRHGRTA
    Non-baseload rates ("≈marginal"): SRNBNOX, SRNBSO2, SRNBCO2, SRNBCH4, SRNBN2O, SRNBC2E, SRNBHG
    Resource mix %:                  SRCLPR, SROLPR, SRGSPR, SRNCPR, SRHYPR, SRBMPR, SRWIPR, SRSOPR, ...

Notes:
- The "non-baseload" rates approximate what a NEW load actually triggers at the margin (the
  next-on generator). Pass use_marginal=True to use these. Default is the all-mix average.
- Social-cost defaults: EPA 2022 SC-GHG ($190/ton CO2), national-average $/ton for NOx/SO2.
  Real damages are location-dependent (EPA COBRA gives county-level multipliers); this
  module uses national averages.
"""
from __future__ import annotations
import argparse
import json
import math
from functools import lru_cache
from pathlib import Path

import openpyxl

ROOT = Path(__file__).parent
EGRID_PATH = ROOT / "echo_bulk" / "egrid2023.xlsx"

# ----- Defaults (override via kwargs) -----
DEFAULT_UTILIZATION = 0.75   # mean IT load fraction over the year (DCs run hot)
DEFAULT_PUE = 1.4            # power usage effectiveness; ~1.15 hyperscale, ~1.6 older

# EPA social costs of pollutants, USD/short ton, 2020 dollars (~2% discount, EPA 2022 SC-GHG +
# BenMAP/COBRA national-average estimates).
SOCIAL_COST_USD_PER_TON = {
    "CO2": 190.0,
    "CH4": 2_000.0,
    "N2O": 56_000.0,
    "NOx": 20_000.0,
    "SO2": 70_000.0,
    # Hg has its own valuation framework; left out of $-rollup but emissions still reported.
}

# Average lb/MWh column codes (total output)
_AVG_RATE_CODES = {
    "NOx": "SRNOXRTA", "SO2": "SRSO2RTA", "CO2": "SRCO2RTA",
    "CH4": "SRCH4RTA", "N2O": "SRN2ORTA", "CO2_equivalent": "SRC2ERTA",
    "Hg": "SRHGRTA",
}
# Non-baseload lb/MWh column codes (≈ marginal)
_MARGINAL_RATE_CODES = {
    "NOx": "SRNBNOX", "SO2": "SRNBSO2", "CO2": "SRNBCO2",
    "CH4": "SRNBCH4", "N2O": "SRNBN2O", "CO2_equivalent": "SRNBC2E",
    "Hg": "SRNBHG",
}
_RESOURCE_MIX_CODES = {
    "coal": "SRCLPR", "oil": "SROLPR", "gas": "SRGSPR", "nuclear": "SRNCPR",
    "hydro": "SRHYPR", "biomass": "SRBMPR", "wind": "SRWIPR", "solar": "SRSOPR",
    "geothermal": "SRGTPR", "other_fossil": "SROFPR",
    "total_nonrenewable": "SRTNPR", "total_renewable": "SRTRPR",
}


# ----- eGRID loading (cached) -----

def _open_egrid() -> openpyxl.Workbook:
    if not EGRID_PATH.exists():
        raise FileNotFoundError(
            f"eGRID workbook not found at {EGRID_PATH}.\n"
            "Download via: curl -o {EGRID_PATH} "
            "https://www.epa.gov/system/files/documents/2025-01/egrid2023_data_rev1.xlsx"
        )
    return openpyxl.load_workbook(str(EGRID_PATH), read_only=True, data_only=True)


@lru_cache(maxsize=1)
def _subregion_table() -> dict:
    """{SUBRGN: {col_code: value, ...}} from SRL23 sheet."""
    wb = _open_egrid()
    ws = wb["SRL23"]
    rows = list(ws.iter_rows(values_only=True))
    # Row 0 = description, Row 1 = column codes, Row 2+ = data
    codes = rows[1]
    out = {}
    for r in rows[2:]:
        if not r or not r[1]:
            continue
        d = {c: v for c, v in zip(codes, r) if c}
        sub = d.get("SUBRGN")
        if isinstance(sub, str):
            out[sub] = d
    return out


@lru_cache(maxsize=1)
def _plant_locations() -> list[tuple]:
    """List of (lat, lon, subrgn) for every eGRID plant — used for nearest-plant lookup."""
    wb = _open_egrid()
    ws = wb["PLNT23"]
    rows = list(ws.iter_rows(values_only=True))
    codes = rows[1]
    idx_lat = codes.index("LAT")
    idx_lon = codes.index("LON")
    idx_sub = codes.index("SUBRGN")
    out = []
    for r in rows[2:]:
        if not r:
            continue
        lat = r[idx_lat]; lon = r[idx_lon]; sub = r[idx_sub]
        if isinstance(lat, (int, float)) and isinstance(lon, (int, float)) and isinstance(sub, str):
            if math.isfinite(lat) and math.isfinite(lon):
                out.append((lat, lon, sub))
    return out


# ----- Geo lookup -----

def _haversine_miles(lat1, lon1, lat2, lon2) -> float:
    R = 3959.0
    lat1r, lat2r = math.radians(lat1), math.radians(lat2)
    dlat, dlon = math.radians(lat2 - lat1), math.radians(lon2 - lon1)
    a = math.sin(dlat / 2) ** 2 + math.cos(lat1r) * math.cos(lat2r) * math.sin(dlon / 2) ** 2
    return 2 * R * math.asin(math.sqrt(a))


def lookup_subregion(lat: float, lon: float) -> dict:
    """Find eGRID subregion via the nearest power plant in the eGRID plant list."""
    plants = _plant_locations()
    best_d = float("inf"); best = None
    for plat, plon, sub in plants:
        d = _haversine_miles(lat, lon, plat, plon)
        if d < best_d:
            best_d = d; best = (plat, plon, sub)
    if best is None:
        raise ValueError(f"No eGRID plant found near lat={lat}, lon={lon}")
    return {
        "subregion_code": best[2],
        "subregion_name": _subregion_table()[best[2]].get("SRNAME"),
        "nearest_plant_lat": best[0],
        "nearest_plant_lon": best[1],
        "nearest_plant_distance_miles": round(best_d, 2),
    }


# ----- Main API -----

def datacenter_pollution(
    lat: float,
    lon: float,
    mw_capacity: float,
    utilization: float = DEFAULT_UTILIZATION,
    pue: float = DEFAULT_PUE,
    use_marginal: bool = False,
) -> dict:
    """Translate a US data center's MW capacity into annual grid pollution via EPA eGRID 2023.

    Args:
        lat, lon: project location (decimal degrees, lon negative for W).
        mw_capacity: IT-side electrical capacity in MW.
        utilization: mean fractional load over the year (default 0.75).
        pue: power usage effectiveness multiplier (default 1.4).
        use_marginal: if True, use eGRID non-baseload rates (≈ marginal new-load impact).
            If False (default), use total-output averages.

    Returns:
        dict with keys:
            input              — echo of inputs and assumptions
            geo                — subregion + nearest-plant info
            annual_MWh         — total electricity drawn from grid per year
            resource_mix       — fraction of subregion generation by fuel type
            emissions          — pollutant -> {lb_per_year, tons_per_year, factor_lb_per_MWh}
            social_cost_usd_per_year         — total monetized externality
            social_cost_breakdown_usd_per_year — per-pollutant $/year
    """
    sub_info = lookup_subregion(lat, lon)
    sub = _subregion_table()[sub_info["subregion_code"]]
    annual_MWh = mw_capacity * 8760.0 * utilization * pue

    rate_codes = _MARGINAL_RATE_CODES if use_marginal else _AVG_RATE_CODES
    emissions = {}
    cost_breakdown = {}
    total_cost = 0.0
    for pollutant, code in rate_codes.items():
        rate_lb_per_MWh = sub.get(code)
        if not isinstance(rate_lb_per_MWh, (int, float)):
            # eGRID often reports '--' for Hg in non-coal regions
            emissions[pollutant] = {
                "lb_per_year": None, "tons_per_year": None,
                "factor_lb_per_MWh": rate_lb_per_MWh,
            }
            continue
        lb = annual_MWh * rate_lb_per_MWh
        tons = lb / 2000.0
        emissions[pollutant] = {
            "lb_per_year": round(lb, 2),
            "tons_per_year": round(tons, 4),
            "factor_lb_per_MWh": rate_lb_per_MWh,
        }
        if pollutant in SOCIAL_COST_USD_PER_TON:
            usd = SOCIAL_COST_USD_PER_TON[pollutant] * tons
            cost_breakdown[pollutant] = round(usd, 0)
            total_cost += usd

    # Resource mix
    mix = {}
    for fuel, code in _RESOURCE_MIX_CODES.items():
        v = sub.get(code)
        mix[fuel] = round(v, 4) if isinstance(v, (int, float)) else None

    return {
        "input": {
            "lat": lat, "lon": lon, "mw_capacity": mw_capacity,
            "utilization": utilization, "pue": pue,
            "rate_basis": "non_baseload (~marginal)" if use_marginal else "total_output_average",
        },
        "geo": sub_info,
        "annual_MWh": round(annual_MWh, 1),
        "resource_mix_pct": mix,
        "emissions": emissions,
        "social_cost_usd_per_year": round(total_cost, 0),
        "social_cost_breakdown_usd_per_year": cost_breakdown,
        "social_cost_assumptions_usd_per_ton": SOCIAL_COST_USD_PER_TON,
    }


# ----- CLI -----

def _cli():
    ap = argparse.ArgumentParser(description="Translate a data center's MW into eGRID-grid pollution.")
    ap.add_argument("--lat", type=float, required=True)
    ap.add_argument("--lon", type=float, required=True)
    ap.add_argument("--mw", type=float, required=True, help="IT-side capacity in MW")
    ap.add_argument("--utilization", type=float, default=DEFAULT_UTILIZATION)
    ap.add_argument("--pue", type=float, default=DEFAULT_PUE)
    ap.add_argument("--marginal", action="store_true",
                    help="Use eGRID non-baseload (~marginal) rates instead of average mix")
    args = ap.parse_args()
    out = datacenter_pollution(
        lat=args.lat, lon=args.lon, mw_capacity=args.mw,
        utilization=args.utilization, pue=args.pue, use_marginal=args.marginal,
    )
    print(json.dumps(out, indent=2))


if __name__ == "__main__":
    _cli()
