"""Stage 1: data ingest.

  1. Convert FracTracker Data_Centers_Database.xlsx to CSV.
  2. Offline-match every FracTracker project against EPA ECHO bulk-download facilities,
     using lat/lon bbox prefilter + haversine + name/address/NAICS scoring.
  3. Attach ICIS-AIR program enrollment dates and ICIS NPDES (water) permit dates per match.

Inputs (must already exist locally):
  ../Data_Centers_Database.xlsx
  echo_bulk/ECHO_EXPORTER.csv               (~2GB, from echo_exporter.zip)
  echo_bulk/ICIS-AIR_PROGRAMS.csv
  echo_bulk/ICIS_PERMITS.csv

Outputs:
  echo_bulk/fractracker.csv                  — FracTracker xlsx as CSV
  echo_bulk/permit_pilot.duckdb              — persistent DB with all joined tables

Run:
  python ingest.py
"""
from __future__ import annotations
import argparse
import csv as csvmod
import json
import re
import sys
from collections import defaultdict
from difflib import SequenceMatcher
from pathlib import Path

import duckdb
from openpyxl import load_workbook

ROOT = Path(__file__).parent
XLSX = ROOT.parent / "Data_Centers_Database.xlsx"
BULK = ROOT / "echo_bulk"
DB = BULK / "permit_pilot.duckdb"

RADIUS_MILES = 1.5
HIGH_CONF_SCORE = 0.5
LAT_PAD = 0.025
LON_PAD = 0.035

# --- NAICS / SIC codes that data centers commonly register under ---
DC_NAICS = {
    "518210", "518111", "518112",
    "541513", "541512", "541519",
    "517311", "517312", "517410", "517919",
    "531120", "236220",
    "221112", "221117", "221118", "221114", "221115",
}
DC_SIC = {"7372", "7374", "7375", "7378", "7379", "4813", "4899",
          "1623", "4911", "4931", "6512"}


# --- Name + address scoring helpers (used by score_candidate) ---
_STRIP_RE = re.compile(
    r"\b(llc|inc|corporation|corp|the|a|an|of|at|by|data center|datacenter|"
    r"data centre|dc|campus|facility|site|hyperscale|colocation|colo|hub|"
    r"project|company|co)\b", re.I,
)
_ADDR_SUFFIX_RE = re.compile(
    r"\b(street|st|road|rd|avenue|ave|boulevard|blvd|highway|hwy|drive|dr|"
    r"lane|ln|court|ct|parkway|pkwy|way|circle|cir|trail|trl|place|pl|terrace|ter)\b\.?", re.I,
)
_ADDR_DIR_RE = re.compile(r"\b(north|south|east|west|nw|ne|sw|se|n|s|e|w)\b\.?", re.I)


def norm_name(n):
    n = (n or "").lower()
    n = _STRIP_RE.sub(" ", n)
    n = re.sub(r"[^a-z0-9 ]+", " ", n)
    return re.sub(r"\s+", " ", n).strip()


def name_score(a, b):
    a, b = norm_name(a), norm_name(b)
    if not a or not b:
        return 0.0
    return SequenceMatcher(None, a, b).ratio()


def norm_addr(a):
    a = (a or "").lower()
    a = _ADDR_SUFFIX_RE.sub(" ", a)
    a = _ADDR_DIR_RE.sub(" ", a)
    a = re.sub(r"[^a-z0-9 ]+", " ", a)
    return re.sub(r"\s+", " ", a).strip()


def addr_score(a, b):
    na, nb = norm_addr(a), norm_addr(b)
    if not na or not nb:
        return 0.0
    base = SequenceMatcher(None, na, nb).ratio()
    ma = re.match(r"^(\d+)\b", na); mb = re.match(r"^(\d+)\b", nb)
    if ma and mb and ma.group(1) == mb.group(1):
        base = min(1.0, base + 0.25)
    return base


def score_candidate(project, fac):
    """Score one ECHO candidate against a FracTracker project. Public for reuse elsewhere."""
    name1 = name_score(project.get("facility_name"), fac.get("FacName"))
    name2 = name_score(project.get("operator_name"), fac.get("FacName"))
    name3 = name_score(project.get("operator_name"), fac.get("FacStreet"))
    best_name = max(name1, name2, name3)
    addr = addr_score(project.get("address"), fac.get("FacStreet"))
    naics_codes = re.split(r"[,\s]+", (fac.get("FacNAICSCodes") or ""))
    sic_codes = re.split(r"[,\s]+", (fac.get("FacSICCodes") or ""))
    naics_hit = any(c in DC_NAICS for c in naics_codes if c)
    sic_hit = any(c in DC_SIC for c in sic_codes if c)
    base = max(best_name, addr)
    score = base + (0.25 if naics_hit else 0.0) + (0.10 if sic_hit else 0.0)
    return {
        "score": round(score, 3),
        "name_score": round(best_name, 3),
        "addr_score": round(addr, 3),
        "naics_hit": naics_hit, "sic_hit": sic_hit,
        "matched_naics": [c for c in naics_codes if c in DC_NAICS],
        "matched_sic": [c for c in sic_codes if c in DC_SIC],
    }


def slugify(s):
    return re.sub(r"[^a-z0-9]+", "_", (s or "").lower()).strip("_")[:80]


# --- Stage 1.1: xlsx → CSV ---
def xlsx_to_csv():
    out = BULK / "fractracker.csv"
    BULK.mkdir(exist_ok=True)
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb["FracTracker Data Centers"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = list(rows[0])
    with out.open("w", newline="") as f:
        w = csvmod.writer(f)
        w.writerow(hdr)
        for r in rows[1:]:
            w.writerow(["" if v is None else v for v in r])
    print(f"[xlsx→csv] {len(rows)-1} rows × {len(hdr)} cols → {out}", file=sys.stderr)


# --- Stage 1.2: bulk match via duckdb ---
def setup(con):
    print("[setup] loading ECHO_EXPORTER (~2GB; ~30s)...", file=sys.stderr)
    con.execute(f"""
      CREATE OR REPLACE TABLE echo AS
      SELECT REGISTRY_ID, FAC_NAME, FAC_STREET, FAC_CITY, FAC_STATE, FAC_ZIP,
             FAC_COUNTY, FAC_FIPS_CODE, FAC_LAT, FAC_LONG,
             FAC_NAICS_CODES, FAC_SIC_CODES,
             AIR_FLAG, NPDES_FLAG, RCRA_FLAG, SDWIS_FLAG, GHG_FLAG,
             AIR_IDS, NPDES_IDS, RCRA_IDS,
             CAA_PERMIT_TYPES, CWA_PERMIT_TYPES, RCRA_PERMIT_TYPES,
             FAC_ACTIVE_FLAG, FAC_MAJOR_FLAG, DFR_URL
      FROM read_csv_auto('{BULK}/ECHO_EXPORTER.csv', sample_size=200000)
      WHERE FAC_LAT IS NOT NULL AND FAC_LONG IS NOT NULL
    """)
    n_echo = con.execute("SELECT COUNT(*) FROM echo").fetchone()[0]
    print(f"[setup] echo: {n_echo:,} facilities", file=sys.stderr)

    con.execute(f"""
      CREATE OR REPLACE TABLE projects AS
      SELECT *, TRY_CAST(lat AS DOUBLE) AS plat, TRY_CAST("long" AS DOUBLE) AS plong
      FROM read_csv_auto('{BULK}/fractracker.csv', sample_size=-1)
    """)
    con.execute(f"""
      CREATE OR REPLACE TABLE air_programs AS
      SELECT * FROM read_csv_auto('{BULK}/ICIS-AIR_PROGRAMS.csv', sample_size=-1)
    """)
    con.execute(f"""
      CREATE OR REPLACE TABLE water_permits AS
      SELECT * FROM read_csv_auto('{BULK}/ICIS_PERMITS.csv', sample_size=-1)
    """)
    n_proj = con.execute("SELECT COUNT(*) FROM projects").fetchone()[0]
    n_air = con.execute("SELECT COUNT(*) FROM air_programs").fetchone()[0]
    n_water = con.execute("SELECT COUNT(*) FROM water_permits").fetchone()[0]
    print(f"[setup] projects={n_proj:,} air_programs={n_air:,} water_permits={n_water:,}", file=sys.stderr)


def find_candidates(con):
    print("[match] bbox prefilter + haversine...", file=sys.stderr)
    con.execute(f"""
      CREATE OR REPLACE TABLE candidates AS
      SELECT
        p.facility_name AS proj_facility_name, p.city AS proj_city, p.state AS proj_state,
        p.county AS proj_county, p.address AS proj_address, p.operator_name AS proj_operator,
        p.status AS proj_status, p.mw AS proj_mw,
        p.plat, p.plong,
        e.REGISTRY_ID, e.FAC_NAME, e.FAC_STREET, e.FAC_CITY, e.FAC_STATE, e.FAC_ZIP,
        e.FAC_COUNTY, e.FAC_FIPS_CODE, e.FAC_LAT, e.FAC_LONG,
        e.FAC_NAICS_CODES, e.FAC_SIC_CODES,
        e.AIR_FLAG, e.NPDES_FLAG, e.RCRA_FLAG, e.SDWIS_FLAG, e.GHG_FLAG,
        e.AIR_IDS, e.NPDES_IDS, e.RCRA_IDS,
        e.CAA_PERMIT_TYPES, e.CWA_PERMIT_TYPES, e.RCRA_PERMIT_TYPES,
        e.FAC_ACTIVE_FLAG, e.FAC_MAJOR_FLAG, e.DFR_URL,
        3959.0 * 2 * asin(sqrt(
          power(sin(radians(e.FAC_LAT - p.plat) / 2), 2) +
          cos(radians(p.plat)) * cos(radians(e.FAC_LAT)) *
          power(sin(radians(e.FAC_LONG - p.plong) / 2), 2)
        )) AS distance_miles
      FROM projects p
      JOIN echo e
        ON e.FAC_LAT  BETWEEN p.plat  - {LAT_PAD} AND p.plat  + {LAT_PAD}
       AND e.FAC_LONG BETWEEN p.plong - {LON_PAD} AND p.plong + {LON_PAD}
      WHERE p.plat IS NOT NULL AND p.plong IS NOT NULL
    """)
    con.execute(f"DELETE FROM candidates WHERE distance_miles > {RADIUS_MILES}")
    n = con.execute("SELECT COUNT(*) FROM candidates").fetchone()[0]
    print(f"[match] {n:,} candidates within {RADIUS_MILES}mi", file=sys.stderr)


def score_and_rank(con):
    print("[score] scoring candidates in Python...", file=sys.stderr)
    rows = con.execute("SELECT * FROM candidates").fetchall()
    cols = [d[0] for d in con.description]
    groups = defaultdict(list)
    for r in rows:
        d = dict(zip(cols, r))
        groups[(d["proj_facility_name"], d["proj_city"], d["proj_state"])].append(d)

    scored = []
    for (pname, pcity, pstate), candidates in groups.items():
        proj = {"facility_name": pname, "city": pcity, "state": pstate,
                "address": candidates[0]["proj_address"],
                "operator_name": candidates[0]["proj_operator"]}
        for c in candidates:
            fac = {"FacName": c["FAC_NAME"], "FacStreet": c["FAC_STREET"],
                   "FacNAICSCodes": c["FAC_NAICS_CODES"], "FacSICCodes": c["FAC_SIC_CODES"]}
            c.update(score_candidate(proj, fac))
        candidates.sort(key=lambda x: (-x["score"], x["distance_miles"]))
        for i, c in enumerate(candidates[:5]):
            c["candidate_rank"] = i + 1
            c["matched_naics"] = ",".join(c.get("matched_naics") or [])
            c["matched_sic"] = ",".join(c.get("matched_sic") or [])
            scored.append(c)
    print(f"[score] {len(scored):,} top-K rows for {len(groups):,} projects", file=sys.stderr)
    if not scored:
        return
    keys = list(scored[0].keys())
    tmp = BULK / "_scored.csv"
    with tmp.open("w", newline="") as f:
        w = csvmod.DictWriter(f, fieldnames=keys)
        w.writeheader()
        for r in scored:
            w.writerow({k: ("" if r.get(k) is None else r.get(k)) for k in keys})
    con.execute("DROP TABLE IF EXISTS scored_matches")
    con.execute(f"CREATE TABLE scored_matches AS SELECT * FROM read_csv_auto('{tmp}', sample_size=-1)")
    tmp.unlink()


def attach_permit_dates(con):
    print("[permit-dates] joining air programs + water permits...", file=sys.stderr)
    con.execute(f"""
      CREATE OR REPLACE TABLE matched_air_programs AS
      SELECT s.proj_facility_name, s.proj_city, s.proj_state, s.candidate_rank,
             s.REGISTRY_ID, s.FAC_NAME,
             ap.PGM_SYS_ID, ap.PROGRAM_CODE, ap.PROGRAM_DESC,
             ap.AIR_OPERATING_STATUS_DESC, ap.BEGIN_DATE, ap.UPDATED_DATE
      FROM scored_matches s
      JOIN air_programs ap ON list_contains(string_split(s.AIR_IDS, ' '), ap.PGM_SYS_ID)
      WHERE s.candidate_rank = 1 AND s.score >= {HIGH_CONF_SCORE}
    """)
    con.execute(f"""
      CREATE OR REPLACE TABLE matched_water_permits AS
      SELECT s.proj_facility_name, s.proj_city, s.proj_state, s.candidate_rank,
             s.REGISTRY_ID, s.FAC_NAME,
             wp.EXTERNAL_PERMIT_NMBR, wp.PERMIT_NAME, wp.PERMIT_TYPE_CODE,
             wp.PERMIT_STATUS_CODE, wp.MAJOR_MINOR_STATUS_FLAG,
             wp.ORIGINAL_ISSUE_DATE, wp.ISSUE_DATE, wp.EFFECTIVE_DATE,
             wp.EXPIRATION_DATE, wp.RETIREMENT_DATE, wp.TERMINATION_DATE,
             wp.ISSUING_AGENCY, wp.AGENCY_TYPE_CODE
      FROM scored_matches s
      JOIN water_permits wp
        ON list_contains(string_split(s.NPDES_IDS, ' '), wp.EXTERNAL_PERMIT_NMBR)
      WHERE s.candidate_rank = 1 AND s.score >= {HIGH_CONF_SCORE}
    """)
    n_air = con.execute("SELECT COUNT(*) FROM matched_air_programs").fetchone()[0]
    n_water = con.execute("SELECT COUNT(*) FROM matched_water_permits").fetchone()[0]
    print(f"[permit-dates] air programs: {n_air:,}, water permits: {n_water:,}", file=sys.stderr)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--skip-xlsx", action="store_true")
    args = ap.parse_args()

    if not args.skip_xlsx:
        xlsx_to_csv()

    if DB.exists():
        DB.unlink()
    print(f"[main] writing duckdb -> {DB}", file=sys.stderr)
    con = duckdb.connect(str(DB))
    setup(con)
    find_candidates(con)
    score_and_rank(con)
    attach_permit_dates(con)
    con.close()
    print(f"[done] DB at {DB} ({DB.stat().st_size/1e6:.0f} MB)", file=sys.stderr)


if __name__ == "__main__":
    main()
