"""Stage 2: LLM permit-event extraction.

Per project:
  1. Build a source-text bundle: FracTracker `other_info` narrative + fetched source URLs +
     DuckDuckGo-discovered news articles filtered by keyword match.
  2. Call Claude Sonnet 4.6 with a forced structured-output tool to extract permit events.
  3. Save extracted/<slug>.json (one file per project).

Public API:
    extract_events(project, fetched_sources=None) -> dict
        Single-project synchronous call; used by both the CLI scaled runner and external code.

CLI:
    python extract.py                        # scaled run, all projects
    python extract.py --limit 5              # 5-project smoke test
    python extract.py --dry-run              # show counts, do nothing

Resumable: skips projects that already have a non-empty extracted file.
"""
from __future__ import annotations
import argparse
import csv as csvmod
import hashlib
import json
import os
import re
import sys
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Any
from urllib.parse import unquote, urlparse

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

ROOT = Path(__file__).parent
XLSX = ROOT.parent / "Data_Centers_Database.xlsx"
EXTRACTED = ROOT / "extracted"
EXTRACTED.mkdir(exist_ok=True)
FETCH_CACHE = ROOT / "fetch_cache"
FETCH_CACHE.mkdir(exist_ok=True)

MODEL = "claude-sonnet-4-6"
UA = "DataCenterPermitTimelinePilot/0.1 (research; ianatroshchenko@gmail.com)"

SKIP_DOMAINS = {
    "youtube.com", "www.youtube.com", "youtu.be",
    "facebook.com", "www.facebook.com", "m.facebook.com",
    "twitter.com", "x.com", "www.x.com",
    "instagram.com", "www.instagram.com",
    "linkedin.com", "www.linkedin.com",
}

# DDG throttle — global serial gate so we don't burst it.
_DDG_LOCK = threading.Lock()
_DDG_LAST = [0.0]

# ----- Schema for the LLM extraction tool -----
EVENT_TYPES = [
    "project_announced", "land_purchased",
    "rezoning_applied", "rezoning_planning_commission_recommendation", "rezoning_council_decision",
    "conditional_use_permit_applied", "conditional_use_permit_decision",
    "special_use_permit_applied", "special_use_permit_decision",
    "variance_applied", "variance_decision",
    "moratorium_enacted", "moratorium_expired",
    "site_plan_applied", "site_plan_approved",
    "building_permit_applied", "building_permit_issued",
    "certificate_of_occupancy_issued",
    "air_permit_applied", "air_permit_issued",
    "water_permit_applied", "water_permit_issued",
    "state_environmental_review_initiated", "state_environmental_review_decision",
    "NEPA_EA_initiated", "NEPA_EA_FONSI",
    "NEPA_EIS_initiated", "NEPA_ROD_issued",
    "USACE_404_applied", "USACE_404_issued",
    "interconnection_request_filed", "interconnection_agreement_executed",
    "power_purchase_agreement_signed",
    "public_hearing_held",
    "lawsuit_filed", "court_ruling", "appeals_court_ruling",
    "project_withdrawn", "project_cancelled", "project_suspended",
    "construction_started", "partial_operational", "fully_operational",
    "other",
]
DECISIONS = ["approved", "denied", "recommended_approval", "recommended_denial",
             "filed", "in_progress", "planned", "withdrawn", "halted", "vacated",
             "granted", "dismissed", "upheld", "remanded", "n_a"]
ACTOR_LEVELS = ["local", "state", "federal", "utility", "court", "developer", "other"]

SYSTEM_PROMPT = f"""You extract structured permit-process events for U.S. data center projects from short narrative text and curated source articles. Your output drives a survival-analysis model of permit timelines, so precision and date accuracy matter more than completeness.

EVENT TAXONOMY — every event must use exactly one event_type from this list:
{json.dumps(EVENT_TYPES, indent=0)}

Use "other" only if no listed type fits, and add an event_subtype string explaining what it was.

DECISION VALUES (set to most precise applicable, "n_a" if not a decision event):
{json.dumps(DECISIONS, indent=0)}

ACTOR_LEVEL is the jurisdictional layer that took the action:
{json.dumps(ACTOR_LEVELS, indent=0)}

DATE RULES — never invent dates. Use ISO format with the precision actually present in the source:
- "2025-11-19" if the source gives a full date
- "2025-11" if only month/year
- "2025" if only year
- Set date_precision to "day" / "month" / "year" accordingly
- If the source uses a relative date and you cannot resolve it from publication date, omit the event

EVERY event MUST have a source_url that came from the input source list AND a source_quote (≤200 chars) — the literal substring of the source supporting the date and event_type. If you cannot produce a verbatim quote, do not emit the event.

Do NOT include events from your own background knowledge that are not supported by the provided sources. The narrative `other_info` field counts as a source — cite it as source_url="other_info" when that is the only place an event is mentioned.

Output a single JSON object via the `report_events` tool."""

EXTRACT_TOOL = {
    "name": "report_events",
    "description": "Report the list of permit-process events extracted from the project material.",
    "input_schema": {
        "type": "object",
        "properties": {
            "facility_name": {"type": "string"},
            "events": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "event_type": {"type": "string", "enum": EVENT_TYPES},
                        "event_subtype": {"type": "string"},
                        "decision": {"type": "string", "enum": DECISIONS},
                        "actor": {"type": "string"},
                        "actor_level": {"type": "string", "enum": ACTOR_LEVELS},
                        "date": {"type": "string"},
                        "date_precision": {"type": "string", "enum": ["day", "month", "year"]},
                        "source_url": {"type": "string"},
                        "source_quote": {"type": "string", "maxLength": 200},
                    },
                    "required": ["event_type", "decision", "actor", "actor_level",
                                 "date", "date_precision", "source_url", "source_quote"],
                },
            },
            "extraction_notes": {"type": "string"},
        },
        "required": ["facility_name", "events", "extraction_notes"],
    },
}


def _api_key() -> str:
    key = os.environ.get("ANTHROPIC_API_KEY")
    if key:
        return key
    keyfile = os.path.expanduser("~/.anthropic_key")
    if os.path.exists(keyfile):
        return open(keyfile).read().strip()
    raise RuntimeError("Set ANTHROPIC_API_KEY or place key in ~/.anthropic_key")


def _slug(s):
    return re.sub(r"[^a-z0-9]+", "_", (s or "").lower()).strip("_")


def project_slug(p):
    return f"{_slug(p.get('facility_name'))[:60]}__{_slug(p.get('city'))[:30]}__{_slug(p.get('state'))[:5]}"


# ----- Single-project extraction (used by external callers) -----
def _user_message(project, fetched_sources):
    parts = [
        f"PROJECT: {project['facility_name']}",
        f"LOCATION: {project['city']}, {project['state']} ({project.get('county', '')})",
        f"STATUS (FracTracker label): {project['status']}",
        f"OPERATOR: {project.get('operator_name') or 'unknown'}",
        f"MW (FracTracker): {project.get('mw') or 'unknown'}",
        "",
        "OTHER_INFO (FracTracker narrative — cite as source_url=\"other_info\"):",
        project.get("other_info") or "(empty)",
        "",
        f"SOURCE URLS (curated by FracTracker, in priority order):",
    ]
    for i, url in enumerate(project.get("sources", []), 1):
        parts.append(f"  [{i}] {url}")
    parts.append("")
    parts.append("FETCHED SOURCE CONTENT:")
    if not fetched_sources:
        parts.append("(none — extract only from OTHER_INFO above)")
    for s in fetched_sources or []:
        parts.append(f"--- {s['url']} ---")
        parts.append(s["text"][:8000])
        parts.append("")
    return "\n".join(parts)


def extract_events(project: dict[str, Any], fetched_sources: list[dict[str, str]] | None = None) -> dict:
    """Call Claude to extract structured events for one project."""
    import anthropic
    client = anthropic.Anthropic(api_key=_api_key())
    response = client.messages.create(
        model=MODEL,
        max_tokens=4096,
        system=[{"type": "text", "text": SYSTEM_PROMPT, "cache_control": {"type": "ephemeral"}}],
        tools=[EXTRACT_TOOL],
        tool_choice={"type": "tool", "name": "report_events"},
        messages=[{"role": "user", "content": _user_message(project, fetched_sources or [])}],
    )
    for block in response.content:
        if block.type == "tool_use" and block.name == "report_events":
            return block.input
    raise RuntimeError("Model did not call report_events tool")


# ----- Source fetching + DDG search -----
session = requests.Session()
session.headers.update({"User-Agent": UA, "Accept": "text/html,application/xhtml+xml",
                        "Accept-Language": "en-US,en;q=0.8"})


def url_cache_path(url):
    h = hashlib.sha1(url.encode()).hexdigest()[:16]
    return FETCH_CACHE / f"{h}.txt"


def fetch_text(url, timeout=12.0):
    if not url or not url.startswith(("http://", "https://")):
        return None
    host = urlparse(url).hostname or ""
    if host.lower() in SKIP_DOMAINS:
        return None
    cache = url_cache_path(url)
    if cache.exists():
        body = cache.read_text()
        return body if body and body != "__BLOCKED__" else None
    try:
        r = session.get(url, timeout=timeout, allow_redirects=True)
        if r.status_code != 200 or not r.text:
            cache.write_text("__BLOCKED__"); return None
        ctype = (r.headers.get("content-type") or "").lower()
        if "html" not in ctype and "text" not in ctype:
            cache.write_text("__BLOCKED__"); return None
        soup = BeautifulSoup(r.text, "lxml")
        for tag in soup(["script", "style", "noscript", "header", "footer", "nav", "aside"]):
            tag.decompose()
        text = re.sub(r"\s+", " ", soup.get_text(separator=" ", strip=True))[:8000]
        cache.write_text(text)
        return text
    except Exception:
        cache.write_text("__BLOCKED__"); return None


def ddg_search(query, max_results=8):
    cache = FETCH_CACHE / f"_ddg_{hashlib.sha1(query.encode()).hexdigest()[:16]}.txt"
    if cache.exists():
        return [u for u in cache.read_text().splitlines() if u]
    with _DDG_LOCK:
        wait = 1.2 - (time.time() - _DDG_LAST[0])
        if wait > 0:
            time.sleep(wait)
        _DDG_LAST[0] = time.time()
    try:
        r = session.post("https://html.duckduckgo.com/html/", data={"q": query}, timeout=15)
        if r.status_code != 200:
            cache.write_text(""); return []
        urls = re.findall(r'href="//duckduckgo\.com/l/\?uddg=([^&"]+)', r.text)
        urls = [unquote(u) for u in urls]
        skip = re.compile(r"(?:wikipedia\.org|amazon\.com|yelp\.com|tiktok\.com)", re.I)
        urls = [u for u in urls if not skip.search(u)][:max_results]
        cache.write_text("\n".join(urls))
        return urls
    except Exception:
        cache.write_text(""); return []


# ----- Scaled runner -----
def load_projects():
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb["FracTracker Data Centers"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = rows[0]
    out = []
    for r in rows[1:]:
        d = dict(zip(hdr, r))
        sources = [d.get(f"info_source_{i}") for i in range(1, 9)]
        d["sources"] = [s for s in sources if s and str(s).strip()]
        out.append(d)
    return out


def needs_extraction(project):
    out_path = EXTRACTED / f"{project_slug(project)}.json"
    if not out_path.exists():
        return True
    try:
        data = json.loads(out_path.read_text())
        return not data.get("events") and not data.get("extraction_notes")
    except Exception:
        return True


def has_useful_input(project):
    return bool((project.get("other_info") or "").strip()) or bool(project.get("sources"))


def process_one(project, max_sources=12, max_search_results=6, use_search=True):
    keywords = [w for w in [project.get("facility_name") or "",
                            project.get("operator_name") or "",
                            project.get("tenant") or "",
                            project.get("city") or ""] if w]
    candidate_urls = []
    seen = set()
    for url in project.get("sources", []):
        if url and url not in seen:
            seen.add(url); candidate_urls.append(url)
    if use_search:
        q_name = project.get("facility_name") or ""
        q_loc = f"{project.get('city') or ''} {project.get('state') or ''}".strip()
        if q_name and len(q_name) > 2:
            try:
                for url in ddg_search(f'"{q_name}" {q_loc} data center', max_results=max_search_results):
                    if url and url not in seen:
                        seen.add(url); candidate_urls.append(url)
            except Exception:
                pass
    fetched = []
    n_fractracker = len(project.get("sources") or [])
    for i, url in enumerate(candidate_urls[:max_sources]):
        text = fetch_text(url)
        if not text:
            continue
        if i >= n_fractracker:
            if not any(k.lower() in (text + " " + url).lower() for k in keywords if len(k) > 4):
                continue
        fetched.append({"url": url, "text": text})
    try:
        result = extract_events(project, fetched_sources=fetched)
        result["_n_sources_fetched"] = len(fetched)
    except Exception as e:
        return {"error": f"{type(e).__name__}: {e}", "facility_name": project["facility_name"]}
    out_path = EXTRACTED / f"{project_slug(project)}.json"
    out_path.write_text(json.dumps(result, indent=2))
    return result


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--limit", type=int, default=0)
    ap.add_argument("--workers", type=int, default=5)
    ap.add_argument("--max-sources", type=int, default=12)
    ap.add_argument("--max-search", type=int, default=6)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    projects = load_projects()
    todo = [p for p in projects if has_useful_input(p) and needs_extraction(p)]
    skipped_no_input = sum(1 for p in projects if not has_useful_input(p))
    skipped_done = len(projects) - len(todo) - skipped_no_input
    if args.limit:
        todo = todo[:args.limit]

    print(f"Total: {len(projects)}  no_input: {skipped_no_input}  already_done: {skipped_done}  todo: {len(todo)}",
          file=sys.stderr)
    if args.dry_run:
        return

    t0 = time.time()
    n_ok = n_err = 0; n_events = 0
    with ThreadPoolExecutor(max_workers=args.workers) as pool:
        futures = {pool.submit(process_one, p, args.max_sources, args.max_search, args.max_search > 0): p
                   for p in todo}
        for i, fut in enumerate(as_completed(futures), 1):
            p = futures[fut]
            try:
                result = fut.result()
                if result is None or "error" in (result or {}):
                    n_err += 1
                    err = result.get("error", "None") if result else "None"
                    print(f"[{i}/{len(todo)}] FAIL {err} {p['facility_name']!r}", file=sys.stderr)
                else:
                    n_ok += 1
                    n = sum(1 for e in (result.get("events") or []) if isinstance(e, dict))
                    n_events += n
                    print(f"[{i}/{len(todo)}] OK {n} events {p['facility_name']!r}", file=sys.stderr)
            except Exception as e:
                n_err += 1
                print(f"[{i}/{len(todo)}] EXC {type(e).__name__}: {e}", file=sys.stderr)
            if i % 25 == 0:
                el = time.time() - t0
                print(f"  ... {i}/{len(todo)} in {el:.0f}s | events={n_events}", file=sys.stderr)

    print(f"\n=== DONE === ok={n_ok} err={n_err} events={n_events} elapsed={time.time()-t0:.0f}s",
          file=sys.stderr)


if __name__ == "__main__":
    main()
