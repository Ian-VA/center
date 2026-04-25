"""Stage 3: build the deliverable CSVs from extracted/*.json + ECHO duckdb.

Runs all stages in order:
  1. wide        — one row per FracTracker project (74 cols), ECHO permit dates joined
  2. timeline    — long-format event CSV (one row per LLM-extracted event)
  3. timelines_by_project — per-project timeline_summary string
  4. filtered    — projects with ≥1 LLM event
  5. mw_filtered — projects with ≥1 LLM event AND parseable MW
  6. durations   — adds 13 timeline-duration columns to (4) and (5)

Outputs into final/:
  data_centers_wide.{csv,parquet}
  permit_timeline_events.{csv,parquet}
  permit_timelines_by_project.csv
  data_centers_with_timelines.{csv,parquet}
  data_centers_with_mw_and_timelines.{csv,parquet}

Run: python build.py
"""
from __future__ import annotations
import csv as csvmod
import json
import re
from collections import defaultdict
from datetime import date
from pathlib import Path

import duckdb
from openpyxl import load_workbook

ROOT = Path(__file__).parent
DB = ROOT / "echo_bulk" / "permit_pilot.duckdb"
EXTRACTED = ROOT / "extracted"
FINAL = ROOT / "final"
FINAL.mkdir(exist_ok=True)

MIN_EVENT_SCORE = 0.5
APPROVAL_DECISIONS = {"approved", "granted", "recommended_approval", "upheld"}
NEGATIVE_DECISIONS = {"denied", "vacated", "halted", "recommended_denial"}
WITHDRAWAL_EVENT_TYPES = {"project_withdrawn", "project_cancelled", "project_suspended"}

INVALID_MW = {"", "None", "Unknown", "unknown", "null", "NA", "n/a"}
_NUM_RE = re.compile(r"\d[\d,]*\.?\d*")


def _slug(s):
    return re.sub(r"[^a-z0-9]+", "_", (s or "").lower()).strip("_")

def project_slug(p):
    return f"{_slug(p.get('facility_name'))[:60]}__{_slug(p.get('city'))[:30]}__{_slug(p.get('state'))[:5]}"


def parse_mw(s):
    if not s or s.strip() in INVALID_MW:
        return None, None, None
    nums = [float(m.group(0).replace(",", "")) for m in _NUM_RE.finditer(s)]
    if not nums:
        return None, None, None
    if len(nums) == 1:
        return nums[0], nums[0], nums[0]
    low, high = min(nums[0], nums[1]), max(nums[0], nums[1])
    return low, high, (low + high) / 2.0


# ===== Stage 1: wide CSV (FracTracker + ECHO permits) =====
def stage_wide():
    print("[wide] building data_centers_wide ...")
    con = duckdb.connect(str(DB))
    con.execute(f"""
      CREATE OR REPLACE VIEW wide AS
      WITH top_match AS (SELECT * FROM scored_matches WHERE candidate_rank = 1),
      air_summary AS (
        SELECT proj_facility_name, proj_city, proj_state,
               COUNT(*) AS n_air_programs,
               MIN(BEGIN_DATE) AS earliest_air_program_begin,
               MAX(BEGIN_DATE) AS latest_air_program_begin,
               STRING_AGG(DISTINCT PROGRAM_CODE, '|') AS air_program_codes,
               STRING_AGG(DISTINCT AIR_OPERATING_STATUS_DESC, '|') AS air_operating_statuses
        FROM matched_air_programs GROUP BY proj_facility_name, proj_city, proj_state
      ),
      water_summary AS (
        SELECT proj_facility_name, proj_city, proj_state,
               COUNT(*) AS n_water_permits,
               MIN(ORIGINAL_ISSUE_DATE) AS earliest_water_orig_issue,
               MIN(ISSUE_DATE) AS earliest_water_issue,
               MAX(ISSUE_DATE) AS latest_water_issue,
               MAX(EXPIRATION_DATE) AS latest_water_expiration,
               MAX(TERMINATION_DATE) AS latest_water_termination,
               STRING_AGG(DISTINCT PERMIT_TYPE_CODE, '|') AS water_permit_types,
               STRING_AGG(DISTINCT PERMIT_STATUS_CODE, '|') AS water_permit_statuses,
               STRING_AGG(DISTINCT EXTERNAL_PERMIT_NMBR, '|') AS water_permit_numbers
        FROM matched_water_permits GROUP BY proj_facility_name, proj_city, proj_state
      )
      SELECT
        p.facility_name, p.address, p.city, p.state, p.zip, p.county,
        p.lat, p."long",
        p.status AS fractracker_status, p.purpose,
        p.operator_name, p.tenant,
        p.mw AS mw_capacity, p.sizerank AS size_rank,
        p.power_source, p.dedicated_power_plant, p.number_of_generators, p.number_of_buildings,
        p.cooling_source, p.cooling_type,
        p.facility_size_sqft, p.property_size_acres, p.project_cost, p.expected_date_online,
        p.community_pushback, p.advocacy_information, p.resistance_status, p.nda,
        s.REGISTRY_ID AS echo_registry_id,
        s.FAC_NAME AS echo_facility_name, s.FAC_STREET AS echo_facility_street,
        s.FAC_CITY AS echo_facility_city, s.FAC_STATE AS echo_facility_state,
        s.FAC_NAICS_CODES AS echo_naics_codes, s.FAC_SIC_CODES AS echo_sic_codes,
        s.AIR_FLAG, s.NPDES_FLAG, s.RCRA_FLAG, s.SDWIS_FLAG, s.GHG_FLAG,
        s.CAA_PERMIT_TYPES, s.CWA_PERMIT_TYPES, s.RCRA_PERMIT_TYPES,
        s.FAC_ACTIVE_FLAG, s.FAC_MAJOR_FLAG,
        s.distance_miles AS echo_distance_miles,
        s.score AS match_score, s.name_score AS match_name_score,
        s.addr_score AS match_addr_score,
        s.naics_hit AS match_naics_hit, s.sic_hit AS match_sic_hit,
        s.matched_naics, s.matched_sic, s.DFR_URL AS echo_dfr_url,
        CASE
          WHEN s.score IS NULL THEN 'none'
          WHEN s.naics_hit AND s.name_score >= 0.4 AND s.distance_miles <= 0.5 THEN 'high'
          WHEN s.naics_hit AND s.distance_miles <= 0.5 THEN 'medium'
          WHEN s.name_score >= 0.6 AND s.distance_miles <= 1.0 THEN 'medium'
          WHEN s.score >= {MIN_EVENT_SCORE} THEN 'low'
          ELSE 'none'
        END AS match_confidence,
        COALESCE(a.n_air_programs, 0) AS n_air_programs,
        a.earliest_air_program_begin, a.latest_air_program_begin,
        a.air_program_codes, a.air_operating_statuses,
        COALESCE(w.n_water_permits, 0) AS n_water_permits,
        w.earliest_water_orig_issue, w.earliest_water_issue, w.latest_water_issue,
        w.latest_water_expiration, w.latest_water_termination,
        w.water_permit_types, w.water_permit_statuses, w.water_permit_numbers,
        NULLIF(LEAST(
          COALESCE(a.earliest_air_program_begin, DATE '9999-12-31'),
          COALESCE(w.earliest_water_orig_issue, DATE '9999-12-31'),
          COALESCE(w.earliest_water_issue, DATE '9999-12-31')
        ), DATE '9999-12-31') AS earliest_known_permit_date
      FROM projects p
      LEFT JOIN top_match s
        ON p.facility_name=s.proj_facility_name AND p.city=s.proj_city AND p.state=s.proj_state
      LEFT JOIN air_summary a
        ON p.facility_name=a.proj_facility_name AND p.city=a.proj_city AND p.state=a.proj_state
      LEFT JOIN water_summary w
        ON p.facility_name=w.proj_facility_name AND p.city=w.proj_city AND p.state=w.proj_state
    """)
    con.execute("CREATE OR REPLACE TABLE final_dataset AS SELECT * FROM wide")

    # Layer in LLM event summaries
    fr = _load_fr()
    by_slug = {project_slug(p): p for p in fr}
    summaries = {}
    for fp in EXTRACTED.glob("*.json"):
        try:
            data = json.loads(fp.read_text())
        except json.JSONDecodeError:
            continue
        proj = by_slug.get(fp.stem)
        if not proj:
            continue
        events = [e for e in (data.get("events") or []) if isinstance(e, dict)]
        dates = sorted([e["date"] for e in events if e.get("date")])
        summaries[(proj["facility_name"], proj["city"], proj["state"])] = {
            "n_llm_events": len(events),
            "earliest_llm_event_date": dates[0] if dates else None,
            "latest_llm_event_date": dates[-1] if dates else None,
            "llm_event_types": "|".join(sorted({e.get("event_type", "") for e in events})),
        }
    con.execute("ALTER TABLE final_dataset ADD COLUMN n_llm_events INTEGER DEFAULT 0")
    con.execute("ALTER TABLE final_dataset ADD COLUMN earliest_llm_event_date VARCHAR")
    con.execute("ALTER TABLE final_dataset ADD COLUMN latest_llm_event_date VARCHAR")
    con.execute("ALTER TABLE final_dataset ADD COLUMN llm_event_types VARCHAR")
    for (n, c, s), e in summaries.items():
        con.execute("""UPDATE final_dataset SET n_llm_events=?, earliest_llm_event_date=?,
                       latest_llm_event_date=?, llm_event_types=?
                       WHERE facility_name=? AND city=? AND state=?""",
                    [e["n_llm_events"], e["earliest_llm_event_date"], e["latest_llm_event_date"],
                     e["llm_event_types"], n, c, s])
    con.execute(f"COPY final_dataset TO '{FINAL/'data_centers_wide.csv'}' (FORMAT CSV, HEADER true)")
    con.execute(f"COPY final_dataset TO '{FINAL/'data_centers_wide.parquet'}' (FORMAT PARQUET)")
    n = con.execute("SELECT COUNT(*) FROM final_dataset").fetchone()[0]
    print(f"  wrote {n} rows × 75 cols")
    con.close()


def _load_fr():
    wb = load_workbook(ROOT.parent / "Data_Centers_Database.xlsx", read_only=True, data_only=True)
    ws = wb["FracTracker Data Centers"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = rows[0]
    return [dict(zip(hdr, r)) for r in rows[1:]]


# ===== Stage 2: long-format events CSV =====
def stage_timeline():
    print("[timeline] building permit_timeline_events + per-project summary ...")
    fr = _load_fr()
    by_slug = {project_slug(p): p for p in fr}
    event_rows = []
    summaries = {}
    for fp in sorted(EXTRACTED.glob("*.json")):
        try:
            data = json.loads(fp.read_text())
        except json.JSONDecodeError:
            continue
        proj = by_slug.get(fp.stem)
        if not proj:
            continue
        events = [e for e in (data.get("events") or []) if isinstance(e, dict)]
        dates = sorted([e["date"] for e in events if e.get("date")])
        summaries[project_slug(proj)] = {
            "n_llm_events": len(events),
            "earliest_llm_event_date": dates[0] if dates else "",
            "latest_llm_event_date": dates[-1] if dates else "",
            "llm_event_types": "|".join(sorted({e.get("event_type", "") for e in events})),
        }
        for e in events:
            event_rows.append({
                "project_facility_name": proj["facility_name"],
                "project_city": proj["city"], "project_state": proj["state"],
                "project_county": proj.get("county"),
                "project_lat": proj.get("lat"), "project_long": proj.get("long"),
                "project_status": proj.get("status"), "project_mw": proj.get("mw"),
                "project_operator": proj.get("operator_name"),
                "event_type": e.get("event_type"), "event_subtype": e.get("event_subtype", ""),
                "decision": e.get("decision"), "actor": e.get("actor"),
                "actor_level": e.get("actor_level"),
                "date": e.get("date"), "date_precision": e.get("date_precision"),
                "source_url": e.get("source_url"),
                "source_quote": (e.get("source_quote", "") or "")[:300],
            })
    if event_rows:
        keys = list(event_rows[0].keys())
        with (FINAL / "permit_timeline_events.csv").open("w", newline="") as f:
            w = csvmod.DictWriter(f, fieldnames=keys); w.writeheader(); w.writerows(event_rows)
        con = duckdb.connect()
        con.execute(f"COPY (SELECT * FROM read_csv_auto('{FINAL/'permit_timeline_events.csv'}')) "
                    f"TO '{FINAL/'permit_timeline_events.parquet'}' (FORMAT PARQUET)")
        print(f"  permit_timeline_events.csv: {len(event_rows)} rows")

    # Per-project summary CSV
    by_proj_events = defaultdict(list)
    for er in event_rows:
        by_proj_events[(er["project_facility_name"], er["project_city"], er["project_state"])].append(er)
    keys = ["facility_name","city","state","county","lat","long","status","mw","operator_name",
            "n_llm_events","earliest_llm_event_date","latest_llm_event_date","llm_event_types",
            "timeline_summary"]
    with (FINAL / "permit_timelines_by_project.csv").open("w", newline="") as f:
        w = csvmod.DictWriter(f, fieldnames=keys); w.writeheader()
        for proj in fr:
            slug = project_slug(proj)
            summ = summaries.get(slug, {})
            evs = by_proj_events.get((proj["facility_name"], proj["city"], proj["state"]), [])
            evs.sort(key=lambda x: (x.get("date") or "9999"))
            tl = "; ".join(f"{e['date']}({e['date_precision']}): {e['event_type']}={e['decision']} by {e['actor']}"
                           for e in evs if e.get("date"))[:1500]
            w.writerow({
                "facility_name": proj["facility_name"], "city": proj["city"], "state": proj["state"],
                "county": proj.get("county"), "lat": proj.get("lat"), "long": proj.get("long"),
                "status": proj.get("status"), "mw": proj.get("mw"),
                "operator_name": proj.get("operator_name"),
                "n_llm_events": summ.get("n_llm_events", 0),
                "earliest_llm_event_date": summ.get("earliest_llm_event_date", ""),
                "latest_llm_event_date": summ.get("latest_llm_event_date", ""),
                "llm_event_types": summ.get("llm_event_types", ""),
                "timeline_summary": tl,
            })
    print(f"  permit_timelines_by_project.csv: {len(fr)} rows")


# ===== Stage 3: filtered subsets (with-events / with-mw) =====
def stage_filtered():
    print("[filtered] data_centers_with_timelines + with_mw_and_timelines ...")
    con = duckdb.connect()
    wide = FINAL / "data_centers_wide.csv"
    tl = FINAL / "permit_timelines_by_project.csv"
    out = FINAL / "data_centers_with_timelines.csv"
    out_pq = FINAL / "data_centers_with_timelines.parquet"
    con.execute(f"""
      CREATE TEMP TABLE w AS SELECT * FROM read_csv_auto('{wide}', sample_size=-1);
      CREATE TEMP TABLE t AS SELECT facility_name, city, state, timeline_summary
                              FROM read_csv_auto('{tl}', sample_size=-1);
      CREATE TABLE filt AS
        SELECT w.*, COALESCE(t.timeline_summary, '') AS timeline_summary
        FROM w LEFT JOIN t USING (facility_name, city, state)
        WHERE w.n_llm_events > 0;
      COPY filt TO '{out}' (FORMAT CSV, HEADER true);
      COPY filt TO '{out_pq}' (FORMAT PARQUET);
    """)
    n = con.execute("SELECT COUNT(*) FROM filt").fetchone()[0]
    print(f"  data_centers_with_timelines.csv: {n} rows")

    # MW-filtered
    rows = list(csvmod.DictReader(out.open()))
    out_mw = FINAL / "data_centers_with_mw_and_timelines.csv"
    out_mw_pq = FINAL / "data_centers_with_mw_and_timelines.parquet"
    kept = []
    for r in rows:
        low, high, mid = parse_mw(r.get("mw_capacity", ""))
        if mid is None:
            continue
        r["mw_low"] = low; r["mw_high"] = high; r["mw_numeric_mid"] = mid
        kept.append(r)
    cols = list(kept[0].keys()) if kept else []
    if kept:
        with out_mw.open("w", newline="") as f:
            w = csvmod.DictWriter(f, fieldnames=cols); w.writeheader(); w.writerows(kept)
        duckdb.connect().execute(
            f"COPY (SELECT * FROM read_csv_auto('{out_mw}', sample_size=-1)) "
            f"TO '{out_mw_pq}' (FORMAT PARQUET)")
        print(f"  data_centers_with_mw_and_timelines.csv: {len(kept)} rows")
    con.close()


# ===== Stage 4: timeline durations =====
def parse_event_date(s, precision):
    if not s or s in ("<UNKNOWN>", ""):
        return None
    parts = s.split("-")
    try:
        if precision == "day" and len(parts) == 3:
            return date(int(parts[0]), int(parts[1]), int(parts[2]))
        if precision == "month" and len(parts) >= 2:
            return date(int(parts[0]), int(parts[1]), 1)
        if precision == "year" and len(parts) >= 1:
            return date(int(parts[0]), 1, 1)
        if len(parts) == 3:
            return date(int(parts[0]), int(parts[1]), int(parts[2]))
        if len(parts) == 2:
            return date(int(parts[0]), int(parts[1]), 1)
        if len(parts) == 1 and len(parts[0]) == 4:
            return date(int(parts[0]), 1, 1)
    except (ValueError, IndexError):
        return None


def compute_durations(events):
    parsed = []
    for e in events:
        d = parse_event_date(e.get("date", ""), e.get("date_precision", ""))
        if d:
            parsed.append((d, e))
    parsed.sort(key=lambda x: x[0])
    if not parsed:
        return {k: "" for k in ("process_start_date","process_start_event_type",
                                "first_approval_date","first_approval_event_type",
                                "first_approval_actor_level","days_to_first_approval",
                                "months_to_first_approval","last_event_date",
                                "total_observed_days","is_approved","is_terminated",
                                "terminal_outcome","is_censored")}
    start_d, start_e = parsed[0]; last_d, _ = parsed[-1]
    fa = next(((d, e) for d, e in parsed if e.get("decision") in APPROVAL_DECISIONS), None)
    is_approved = fa is not None
    is_negated = any(e.get("decision") in NEGATIVE_DECISIONS for _, e in parsed)
    is_withdrawn = any(e.get("event_type") in WITHDRAWAL_EVENT_TYPES for _, e in parsed)
    if is_negated:
        terminal = "denied"; is_term = True
    elif is_withdrawn:
        terminal = "withdrawn"; is_term = True
    elif is_approved:
        terminal = "approved"; is_term = False
    else:
        terminal = "pending"; is_term = False
    out = {
        "process_start_date": start_d.isoformat(),
        "process_start_event_type": start_e.get("event_type", ""),
        "first_approval_date": "", "first_approval_event_type": "", "first_approval_actor_level": "",
        "days_to_first_approval": "", "months_to_first_approval": "",
        "last_event_date": last_d.isoformat(),
        "total_observed_days": (last_d - start_d).days,
        "is_approved": is_approved, "is_terminated": is_term,
        "terminal_outcome": terminal,
        "is_censored": (not is_approved) and (not is_negated) and (not is_withdrawn),
    }
    if fa:
        d, e = fa
        out["first_approval_date"] = d.isoformat()
        out["first_approval_event_type"] = e.get("event_type", "")
        out["first_approval_actor_level"] = e.get("actor_level", "")
        delta = max(0, (d - start_d).days)
        out["days_to_first_approval"] = delta
        out["months_to_first_approval"] = round(delta / 30.4375, 1)
    return out


def stage_durations():
    print("[durations] adding 13 timeline-duration columns ...")
    events = list(csvmod.DictReader((FINAL / "permit_timeline_events.csv").open()))
    by_proj = defaultdict(list)
    for e in events:
        by_proj[(e["project_facility_name"], e["project_city"], e["project_state"])].append(e)
    durs = {k: compute_durations(evs) for k, evs in by_proj.items()}
    NEW_COLS = list(next(iter(durs.values())).keys())
    for src in [FINAL / "data_centers_with_timelines.csv", FINAL / "data_centers_with_mw_and_timelines.csv"]:
        if not src.exists():
            continue
        rows = list(csvmod.DictReader(src.open()))
        cols = list(rows[0].keys()) + [c for c in NEW_COLS if c not in rows[0]]
        for r in rows:
            d = durs.get((r["facility_name"], r["city"], r["state"]), {})
            for c in NEW_COLS:
                r[c] = d.get(c, "")
        with src.open("w", newline="") as f:
            w = csvmod.DictWriter(f, fieldnames=cols); w.writeheader(); w.writerows(rows)
        duckdb.connect().execute(
            f"COPY (SELECT * FROM read_csv_auto('{src}', sample_size=-1)) "
            f"TO '{src.with_suffix('.parquet')}' (FORMAT PARQUET)")
        print(f"  updated {src.name}")


def main():
    stage_wide()
    stage_timeline()
    stage_filtered()
    stage_durations()
    print("\n[done] all stages complete; outputs in final/")


if __name__ == "__main__":
    main()
